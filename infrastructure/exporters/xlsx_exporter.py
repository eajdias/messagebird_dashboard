"""XLSX Exporter — generates conversation XLSX files with openpyxl."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_BLUE = "1A3A5C"
HEADER_WHITE = "FFFFFF"
ALT_ROW = "F3F6FA"
BORDER_COLOR = "D0D5DD"

COLUMNS = [
    ("Contato", "contact", 24),
    ("Telefone", "phone", 18),
    ("Agente", "agent", 22),
    ("Departamento", "department", 18),
    ("Mensagens", "msg_count", 11),
    ("Nota (Agente)", "rating", 14),
    ("NPS", "nps", 8),
    ("ART (min)", "art_minutes", 12),
    ("Data", "start_time", 14),
]

RIGHT_ALIGN_COLS = {"msg_count", "rating", "nps", "art_minutes"}


def _normalize_row(r: dict[str, Any]) -> dict[str, Any]:
    from domain import constants

    dept_id = r.get("cnvs_dept")
    dept_label = ""
    if "department" in r and r["department"] and isinstance(r["department"], str) and not r["department"].isdigit():
        dept_label = r["department"]
    elif dept_id is not None:
        dept_label = constants.resolve_dept(int(dept_id) if isinstance(dept_id, str) else dept_id)

    return {
        "contact": r.get("cnts_name") or r.get("contact", ""),
        "phone": r.get("cnts_phone") or r.get("phone", ""),
        "agent": r.get("agnt_name") or r.get("agent", ""),
        "department": dept_label,
        "msg_count": r.get("cnvs_msgcount") or r.get("msg_count", 0),
        "rating": r.get("cnvs_rating_agent") if r.get("cnvs_rating_agent") is not None else r.get("rating"),
        "nps": r.get("cnvs_rating_nps") if r.get("cnvs_rating_nps") is not None else r.get("nps"),
        "art_minutes": r.get("cnvs_art_minutes") if r.get("cnvs_art_minutes") is not None else r.get("art_minutes"),
        "start_time": _fmt_ts(r.get("cnvs_created") or r.get("start_time")),
    }


def _fmt_ts(val: Any) -> str:
    if not val:
        return ""
    try:
        from datetime import datetime

        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y %H:%M")
        return str(val)[:19]
    except Exception:
        return str(val)


class XLSXExporter:
    def generate(self, rows: list[dict[str, Any]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Conversas"

        border_style = Side(style="thin", color=BORDER_COLOR)
        all_borders = Border(
            top=border_style,
            bottom=border_style,
            left=border_style,
            right=border_style,
        )
        header_font = Font(bold=True, color=HEADER_WHITE, size=11)
        header_fill = PatternFill(patternType="solid", fgColor=HEADER_BLUE)
        alt_fill = PatternFill(patternType="solid", fgColor=ALT_ROW)

        for col_idx, (label, _, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
            cell.border = all_borders
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 28

        for row_idx, r in enumerate(rows, 2):
            n = _normalize_row(r)
            for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
                value = n.get(key)
                if key in RIGHT_ALIGN_COLS and value is not None:
                    if key == "nps":
                        value = int(round(float(value), 0))
                    elif key == "art_minutes":
                        value = round(float(value), 1)
                    elif key == "rating":
                        value = int(value) if value == int(value) else value
                if value is None:
                    value = ""

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = all_borders

                if key in RIGHT_ALIGN_COLS:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center")

                if row_idx % 2 == 0:
                    cell.fill = alt_fill

            ws.row_dimensions[row_idx].height = 22

        last_row = ws.max_row
        last_col = get_column_letter(len(COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
