"""Dashboard Export Service — generates multi-sheet Excel from dashboard data."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_BLUE = "1A3A5C"
HEADER_WHITE = "FFFFFF"
ALT_ROW = "F3F6FA"
BORDER_COLOR = "D0D5DD"


def _sheet_style(ws) -> None:
    border = Side(style="thin", color=BORDER_COLOR)
    all_borders = Border(top=border, bottom=border, left=border, right=border)
    header_font = Font(bold=True, color=HEADER_WHITE, size=11)
    header_fill = PatternFill(patternType="solid", fgColor=HEADER_BLUE)
    alt_fill = PatternFill(patternType="solid", fgColor=ALT_ROW)

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
        cell.border = all_borders
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    ws.row_dimensions[1].height = 28

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = all_borders
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 22

    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    ws.freeze_panes = "A2"


class DashboardExportService:
    def generate(self, processed: list[Any], sections: set[str]) -> bytes:
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        if "summary" in sections:
            self._add_summary(wb, processed)
        if "evolution" in sections:
            self._add_evolution(wb, processed)
        if "quality" in sections:
            self._add_quality(wb, processed)
        if "agents" in sections:
            self._add_agents(wb, processed)
        if "departments" in sections:
            self._add_departments(wb, processed)
        if "motives" in sections:
            self._add_motives(wb, processed)
        if "heatmap" in sections:
            self._add_heatmap(wb, processed)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _add_summary(self, wb: Workbook, processed: list[Any]) -> None:
        ws = wb.create_sheet("Resumo KPIs")
        from application.services.report_aggregator import ReportAggregator

        agg = ReportAggregator()
        stats = agg.aggregate_statistics(processed)

        ws.append(["Indicador", "Valor"])
        ws.append(["Total de conversas", stats.get("total_chats", 0)])
        ws.append(["Total de mensagens", stats.get("total_msgs", 0)])
        ws.append(["NPS (real)", self._n(stats.get("real_nps"))])
        ws.append(["ART médio (min)", self._n(stats.get("avg_art"))])
        ws.append(["Duração média (min)", self._n(stats.get("avg_duration"))])
        ws.append(["SLA compliance", self._pct(stats.get("sla_compliance"))])
        ws.append(["Rating médio", self._n(stats.get("avg_rating"))])
        ws.append(["Elogios", stats.get("compliments", 0)])
        ws.append(["Negativos", stats.get("negatives", 0)])
        ws.append(["Contatos únicos", stats.get("unique_clients", 0)])
        ws.append(["Retornantes", stats.get("returners", 0)])
        _sheet_style(ws)

    def _add_quality(self, wb: Workbook, processed: list[Any]) -> None:
        ws = wb.create_sheet("Qualidade e NPS")
        from application.services.sub_aggregators import RatingAggregator

        dist = RatingAggregator().aggregate_distributions(processed)

        # NPS breakdown
        nps_dist = dist.get("nps_distribution", {})
        ws.append(["Métrica", "Valor", "Percentual"])
        total_nps = (
            int(nps_dist.get("promoters", 0)) + int(nps_dist.get("passives", 0)) + int(nps_dist.get("detractors", 0))
        )
        if total_nps > 0:
            ws.append(
                [
                    "Promotores",
                    int(nps_dist.get("promoters", 0)),
                    self._pct(int(nps_dist.get("promoters", 0)) / total_nps * 100),
                ]
            )
            ws.append(
                [
                    "Neutros",
                    int(nps_dist.get("passives", 0)),
                    self._pct(int(nps_dist.get("passives", 0)) / total_nps * 100),
                ]
            )
            ws.append(
                [
                    "Detratores",
                    int(nps_dist.get("detractors", 0)),
                    self._pct(int(nps_dist.get("detractors", 0)) / total_nps * 100),
                ]
            )
        else:
            ws.append(["Sem dados", "", ""])

        # Rating 1-5
        rating_counts = dist.get("rating_distribution", {})
        ws.append([])
        ws.append(["Nota", "Quantidade", "Percentual"])
        rating_total = sum(int(rating_counts.get(str(i), 0)) for i in range(1, 6))
        if rating_total > 0:
            for i in range(1, 6):
                c = int(rating_counts.get(str(i), 0))
                ws.append([str(i), c, self._pct(c / rating_total * 100)])
        else:
            ws.append(["Sem dados", "", ""])

        # NPS 1-10
        ws.append([])
        ws.append(["NPS", "Quantidade", "Percentual"])
        nps_raw: dict[str, int] = {str(i): 0 for i in range(1, 11)}
        for p in processed:
            if p.nps is not None and 1 <= p.nps <= 10:
                nps_raw[str(int(p.nps))] += 1
        nps_total_raw = sum(nps_raw.values())
        if nps_total_raw > 0:
            for i in range(1, 11):
                ws.append([str(i), nps_raw[str(i)], self._pct(nps_raw[str(i)] / nps_total_raw * 100)])
        else:
            ws.append(["Sem dados", "", ""])
        _sheet_style(ws)

    def _add_departments(self, wb: Workbook, processed: list[Any]) -> None:
        ws = wb.create_sheet("Departamentos")

        by_dept: dict[str, list[Any]] = {}
        for p in processed:
            by_dept.setdefault(p.dept_label, []).append(p)

        ws.append(["Departamento", "Chats", "NPS Real", "ART Médio (min)", "Rating Médio", "Retornantes"])
        for dept, items in sorted(by_dept.items()):
            chats = [p for p in items]
            ratings = [p.rating for p in chats if p.rating is not None]
            nps_vals = [p.nps for p in chats if p.nps is not None]
            arts = [p.art_min for p in chats if p.art_min is not None]
            from domain.services.metrics_calculator import MetricsCalculator

            nps = MetricsCalculator.calculate_nps(nps_vals) if nps_vals else None
            ws.append(
                [
                    dept,
                    len(chats),
                    self._n(nps),
                    self._n(round(sum(arts) / len(arts), 1)) if arts else "",
                    self._n(round(sum(ratings) / len(ratings), 1)) if ratings else "",
                    sum(1 for p in chats if p.contact_id),
                ]
            )
        _sheet_style(ws)

    def _add_agents(self, wb: Workbook, processed: list[Any]) -> None:
        ws = wb.create_sheet("Ranking de Agentes")
        ws.append(
            [
                "Agente",
                "Chats",
                "NPS",
                "Rating Médio",
                "ART Médio (min)",
                "Elogios",
                "Negativos",
                "Contatos Únicos",
                "SLA %",
            ]
        )

        by_agent: dict[str, list[Any]] = {}
        for p in processed:
            agent = p.agent or "Desconhecido"
            by_agent.setdefault(agent, []).append(p)

        rows = []
        for agent, items in sorted(by_agent.items()):
            ratings = [p.rating for p in items if p.rating is not None]
            nps_vals = [p.nps for p in items if p.nps is not None]
            arts = [p.art_min for p in items if p.art_min is not None]
            from domain.services.metrics_calculator import MetricsCalculator

            nps = MetricsCalculator.calculate_nps(nps_vals) if nps_vals else None
            rows.append(
                (
                    len(items),
                    agent,
                    self._n(nps),
                    self._n(round(sum(ratings) / len(ratings), 1)) if ratings else "",
                    self._n(round(sum(arts) / len(arts), 1)) if arts else "",
                    sum(1 for p in items if p.is_compliment),
                    sum(1 for p in items if p.is_negative),
                    len({p.contact_id for p in items if p.contact_id}),
                    "",
                )
            )

        for r in sorted(rows, reverse=True):
            ws.append(list(r[1:]))
        _sheet_style(ws)

    def _add_motives(self, wb: Workbook, processed: list[Any]) -> None:
        ws = wb.create_sheet("Motivos e Ocorrências")
        from collections import Counter

        motives = Counter(p.contact_reason for p in processed if p.contact_reason)
        occurrences = Counter(p.occurrence for p in processed if p.occurrence)

        ws.append(["Motivo de Contato", "Quantidade"])
        for label, count in motives.most_common():
            ws.append([label, count])
        ws.append([])
        ws.append([])
        ws.append(["Ocorrência", "Quantidade"])
        for label, count in occurrences.most_common():
            ws.append([label, count])
        _sheet_style(ws)

    def _add_evolution(self, wb: Workbook, processed: list[Any]) -> None:
        ws = wb.create_sheet("Evolução Mensal")
        from collections import defaultdict

        by_month: dict[str, dict[str, Any]] = defaultdict(lambda: {"chats": 0, "nps": [], "art": [], "rating": []})
        for p in processed:
            raw = getattr(p, "raw_created", "")
            raw_obj: Any = raw
            if raw and isinstance(raw_obj, str) and len(raw_obj) >= 7:
                key = raw_obj[:7]
            elif hasattr(raw_obj, "strftime"):
                key = raw_obj.strftime("%Y-%m")
            else:
                continue
            by_month[key]["chats"] += 1
            if p.nps is not None:
                by_month[key]["nps"].append(p.nps)
            if p.art_min is not None:
                by_month[key]["art"].append(p.art_min)
            if p.rating is not None:
                by_month[key]["rating"].append(p.rating)

        ws.append(["Mês", "Chats", "NPS", "ART Médio (min)", "Rating Médio"])
        from domain.services.metrics_calculator import MetricsCalculator

        for month in sorted(by_month.keys()):
            d = by_month[month]
            nps = MetricsCalculator.calculate_nps(d["nps"]) if d["nps"] else None
            art = round(sum(d["art"]) / len(d["art"]), 1) if d["art"] else None
            rating = round(sum(d["rating"]) / len(d["rating"]), 1) if d["rating"] else None
            ws.append([month, d["chats"], self._n(nps), self._n(art), self._n(rating)])
        _sheet_style(ws)

    def _add_heatmap(self, wb: Workbook, processed: list[Any]) -> None:
        ws = wb.create_sheet("Heatmap")
        from collections import Counter

        cells = Counter()
        for p in processed:
            raw = getattr(p, "raw_created", "")
            raw_obj: Any = raw
            if isinstance(raw_obj, str) and len(raw_obj) >= 13:
                hour = int(raw_obj[11:13]) if raw_obj[11:13].isdigit() else 0
                day = int(raw_obj[8:10]) if raw_obj[8:10].isdigit() else 0
            elif hasattr(raw_obj, "strftime"):
                hour = raw_obj.hour
                day = raw_obj.day
            else:
                continue
            cells[(day, hour)] += 1

        days = sorted({d for d, _ in cells})
        hours = sorted({h for _, h in cells})
        ws.append(["Hora"] + [f"Dia {d}" for d in days])
        for h in hours:
            row_data = [f"{h:02d}h"]
            for d in days:
                row_data.append(cells.get((d, h), 0))
            ws.append(row_data)
        _sheet_style(ws)

    def _n(self, val: Any) -> float | str:
        if val is None:
            return ""
        if isinstance(val, (int, float)):
            return round(val, 1)
        return str(val)

    def _pct(self, val: Any) -> str:
        if val is None:
            return ""
        try:
            return f"{float(val):.1f}%"
        except TypeError, ValueError:
            return str(val)
