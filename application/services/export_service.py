"""Export Service — generates CSV, XLSX, and ZIP (OS) files from conversation data."""

from __future__ import annotations

import io
import logging
import os
import uuid
import zipfile
from collections import defaultdict
from datetime import datetime
from typing import Any

from domain import constants
from infrastructure.exporters.csv_exporter import CSVExporter
from infrastructure.exporters.pdf_exporter import PDFExporter
from infrastructure.exporters.xlsx_exporter import XLSXExporter

logger = logging.getLogger("m_bird.export_service")

EXPORT_DIRS: dict[str, str] = {
    "csv": "exports/csv",
    "xlsx": "exports/xlsx",
    "pdf_zip": "exports/zip",
}


def _fmt_ts(val: Any) -> str:
    if not val:
        return ""
    if isinstance(val, str):
        return val[:19]
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val)[:19]


class ExportService:
    def __init__(self, reports_dir: str):
        self._reports_dir = reports_dir
        self._csv = CSVExporter()
        self._xlsx = XLSXExporter()
        self._pdf = PDFExporter()

    def save_csv(self, rows: list[dict[str, Any]], start_date: str, end_date: str, created_by: str) -> dict[str, Any]:
        data = self._csv.generate(rows)
        return self._save_file("csv", "csv", data, start_date, end_date, rows, created_by)

    def save_xlsx(self, rows: list[dict[str, Any]], start_date: str, end_date: str, created_by: str) -> dict[str, Any]:
        data = self._xlsx.generate(rows)
        return self._save_file("xlsx", "xlsx", data, start_date, end_date, rows, created_by)

    def save_zip_os(
        self,
        rows: list[dict[str, Any]],
        msgs_map: dict[int, list[dict[str, Any]]],
        start_date: str,
        end_date: str,
        created_by: str,
    ) -> dict[str, Any]:
        zip_buf = io.BytesIO()
        count = 0
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for r in rows:
                cnvs_id = r.get("cnvs_id") or r.get("id")
                if not cnvs_id:
                    continue
                cnvs_id = int(cnvs_id)
                detail = self._row_to_detail(r)
                msgs = msgs_map.get(cnvs_id, [])
                pdf_bytes = self._pdf.generate_single_os_pdf_bytes(detail, msgs)
                protocol = detail.get("cnvs_bird") or str(cnvs_id)
                zf.writestr(f"OS_{protocol}.pdf", pdf_bytes)
                count += 1

        zip_buf.seek(0)
        data = zip_buf.getvalue()
        return self._save_file("pdf_zip", "zip", data, start_date, end_date, [{"_count": count}], created_by)

    def save_returners_report(
        self,
        rows: list[dict[str, Any]],
        start_date: str,
        end_date: str,
        created_by: str,
        format: str = "csv",
    ) -> dict[str, Any]:
        """Aggregate conversations by contact_id for the returners report."""
        by_contact: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for r in rows:
            cid = r.get("cnts_id") or r.get("contact_id", 0)
            if cid:
                by_contact[cid].append(r)

        agg_rows: list[dict[str, Any]] = []
        for _cid, convs in sorted(by_contact.items()):
            first = convs[-1]
            last = convs[0]
            contact = first.get("cnts_name") or first.get("contact", "")
            phone = first.get("cnts_phone") or first.get("phone", "")
            visit_count = len(convs)
            agents = sorted({str(c.get("agnt_name") or c.get("agent", "Desconhecido")) for c in convs})
            departments = sorted(
                {constants.resolve_dept(int(c.get("cnvs_dept"))) for c in convs if c.get("cnvs_dept") is not None}
            )
            channels = sorted(
                {constants.resolve_channel(str(c.get("cnvs_channel"))) for c in convs if c.get("cnvs_channel")}
            )
            ratings = [c.get("cnvs_rating_agent") for c in convs if c.get("cnvs_rating_agent") is not None]
            nps_vals = [c.get("cnvs_rating_nps") for c in convs if c.get("cnvs_rating_nps") is not None]
            arts = [c.get("cnvs_art_minutes") for c in convs if c.get("cnvs_art_minutes") is not None]

            # Count distinct days
            days = set()
            for c in convs:
                val = c.get("cnvs_created")
                if val:
                    if isinstance(val, str):
                        days.add(val[:10])
                    elif hasattr(val, "strftime"):
                        days.add(val.strftime("%Y-%m-%d"))
                    else:
                        days.add(str(val)[:10])

            agg_rows.append(
                {
                    "contact": contact,
                    "phone": phone,
                    "visitas": visit_count,
                    "dias_distintos": len(days),
                    "agentes": ", ".join(agents),
                    "departamentos": ", ".join(map(str, departments)),
                    "canais": ", ".join(map(str, channels)),
                    "nota_media": round(sum(ratings) / len(ratings), 1) if ratings else "",
                    "nps_medio": round(sum(nps_vals) / len(nps_vals), 0) if nps_vals else "",
                    "art_medio_min": round(sum(arts) / len(arts), 1) if arts else "",
                    "primeira_visita": _fmt_ts(last.get("cnvs_created", "")),
                    "ultima_visita": _fmt_ts(first.get("cnvs_created", "")),
                    "total_msgs": sum(c.get("cnvs_msgcount", 0) for c in convs),
                }
            )

        prefix = "retornantes" if format == "csv" else "xlsx"
        ext = "csv" if format == "csv" else "xlsx"
        if format == "csv":
            data = self._csv_aggregated(agg_rows, RETURNER_COLUMNS)
        else:
            data = self._xlsx_aggregated(agg_rows, RETURNER_COLUMNS, "Retornantes")

        return self._save_file(prefix, ext, data, start_date, end_date, agg_rows, created_by)

    def save_art_high_report(
        self,
        rows: list[dict[str, Any]],
        start_date: str,
        end_date: str,
        created_by: str,
        format: str = "csv",
    ) -> dict[str, Any]:
        """Save ART > threshold report (rows already filtered by the query)."""
        prefix = "art_high" if format == "csv" else "xlsx"
        ext = "csv" if format == "csv" else "xlsx"
        data = self._csv.generate(rows) if format == "csv" else self._xlsx.generate(rows)

        return self._save_file(prefix, ext, data, start_date, end_date, rows, created_by)

    def _csv_aggregated(self, rows: list[dict[str, Any]], columns: list[tuple[str, str]]) -> bytes:
        import csv

        buf = io.StringIO()
        buf.write("\ufeff")
        fieldnames = [c[0] for c in columns]
        labels = [c[1] for c in columns]
        writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
        writer.writerow(dict(zip(fieldnames, labels, strict=True)))
        for r in rows:
            writer.writerow(r)
        return buf.getvalue().encode("utf-8")

    def _xlsx_aggregated(self, rows: list[dict[str, Any]], columns: list[tuple[str, str]], sheet_name: str) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = sheet_name

        border_style = Side(style="thin", color="D0D5DD")
        all_borders = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(patternType="solid", fgColor="1A3A5C")
        alt_fill = PatternFill(patternType="solid", fgColor="F3F6FA")

        for col_idx, (_, label) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
            cell.border = all_borders
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        ws.row_dimensions[1].height = 28

        for row_idx, r in enumerate(rows, 2):
            for col_idx, (key, _) in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=r.get(key, ""))
                cell.border = all_borders
                cell.alignment = Alignment(vertical="center")
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
            ws.row_dimensions[row_idx].height = 22

        last_row = ws.max_row
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _save_file(
        self,
        format: str,
        ext: str,
        data: bytes,
        start_date: str,
        end_date: str,
        rows: list[dict[str, Any]],
        created_by: str,
    ) -> dict[str, Any]:
        report_id = f"{format}_{uuid.uuid4().hex[:8]}"
        safe_ts = datetime.now().strftime("%H%M%S")
        filename = f"conversas_{start_date}_{end_date}_{safe_ts}.{ext}"

        rel_dir = EXPORT_DIRS.get(format, "exports")
        abs_dir = os.path.join(self._reports_dir, rel_dir)
        os.makedirs(abs_dir, exist_ok=True)

        filepath = os.path.join(abs_dir, filename)
        with open(filepath, "wb") as f:
            f.write(data)

        size_bytes = os.path.getsize(filepath)

        return {
            "report_id": report_id,
            "type": "export",
            "format": format,
            "start_date": start_date,
            "end_date": end_date,
            "filename": filename,
            "path": os.path.relpath(filepath, self._reports_dir),
            "size_bytes": size_bytes,
            "record_count": len(rows),
            "created_by": created_by,
            "created_at": datetime.now().isoformat(),
        }

    def _row_to_detail(self, r: dict[str, Any]) -> dict[str, Any]:
        return {
            "cnvs_id": r.get("cnvs_id") or r.get("id", 0),
            "cnvs_created": r.get("cnvs_created") or r.get("start_time", ""),
            "cnvs_updated": r.get("cnvs_updated") or r.get("end_time", ""),
            "cnvs_status": r.get("cnvs_status", ""),
            "cnvs_dept": r.get("cnvs_dept"),
            "cnvs_rating_agent": r.get("cnvs_rating_agent"),
            "cnvs_rating_nps": r.get("cnvs_rating_nps"),
            "cnvs_msgcount": r.get("cnvs_msgcount") or r.get("msg_count", 0),
            "cnvs_reopened_count": r.get("cnvs_reopened_count", 0),
            "cnvs_channel": r.get("cnvs_channel", ""),
            "cnvs_description": r.get("cnvs_description", ""),
            "cnvs_bird": r.get("cnvs_bird", ""),
            "cnvs_tax_id": r.get("cnvs_tax_id", ""),
            "cnvs_software": r.get("cnvs_software", ""),
            "cnvs_contact_reason": r.get("cnvs_contact_reason"),
            "cnvs_occurrence": r.get("cnvs_occurrence"),
            "agnt_name": r.get("agnt_name") or r.get("agent", ""),
            "cnts_name": r.get("cnts_name") or r.get("contact", ""),
            "cnts_phone": r.get("cnts_phone") or r.get("phone", ""),
        }


RETURNER_COLUMNS: list[tuple[str, str]] = [
    ("contact", "Contato"),
    ("phone", "Telefone"),
    ("visitas", "Visitas"),
    ("dias_distintos", "Dias distintos"),
    ("agentes", "Agentes"),
    ("departamentos", "Departamentos"),
    ("canais", "Canais"),
    ("nota_media", "Nota média"),
    ("nps_medio", "NPS médio"),
    ("art_medio_min", "ART médio (min)"),
    ("total_msgs", "Total msgs"),
    ("primeira_visita", "1ª visita"),
    ("ultima_visita", "Última visita"),
]
